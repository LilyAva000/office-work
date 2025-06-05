import re
from copy import copy
from openpyxl import load_workbook

class ExcelTemplateFiller:
    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active
        self.sample_row_idx, self.template_cells = self._find_sample_row_and_cells()

    @staticmethod
    def get_value_by_key(expr, data):
        expr = expr.strip("{} ").replace("][", "].[")
        parts = re.split(r"\[|\]", expr)
        key = parts[0]
        value = data.get(key)
        for part in parts[1:]:
            part = part.strip(" '\"")
            if part:
                if isinstance(value, dict):
                    value = value.get(part)
                elif isinstance(value, list) and part.isdigit():
                    value = value[int(part)]
        return value

    def _find_sample_row_and_cells(self):
        sample_row_idx = None
        for i, row in enumerate(self.ws.iter_rows(min_row=1, max_row=self.ws.max_row), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "{{" in cell.value:
                    sample_row_idx = i
                    break
            if sample_row_idx:
                break
        if sample_row_idx is None:
            raise RuntimeError("未找到包含占位符的样板行，请确认模板中包含如 {{姓名}} 这种变量行。")
        template_cells = [cell for cell in self.ws[sample_row_idx]]
        return sample_row_idx, template_cells

    def fill_row(self, row_idx, data):
        for col_idx, template_cell in enumerate(self.template_cells, 1):
            new_cell = self.ws.cell(row=row_idx, column=col_idx)
            # 样式复制
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.border = copy(template_cell.border)
                new_cell.fill = copy(template_cell.fill)
                new_cell.number_format = copy(template_cell.number_format)
                new_cell.protection = copy(template_cell.protection)
                new_cell.alignment = copy(template_cell.alignment)
            # 内容替换
            cell_val = template_cell.value
            if cell_val and "{{" in str(cell_val):
                matches = re.findall(r"{{([^{}]+)}}", str(cell_val))
                for m in matches:
                    expr = "{{" + m + "}}"
                    val = self.get_value_by_key(expr, data)
                    cell_val = cell_val.replace(expr, str(val) if val is not None else "")
            new_cell.value = cell_val

    def fill(self, info):
        # 删除样板行
        self.ws.delete_rows(self.sample_row_idx, 1)
        if isinstance(info, list):
            for idx, item in enumerate(info):
                self.ws.insert_rows(self.sample_row_idx + idx)
                self.fill_row(self.sample_row_idx + idx, item)
        elif isinstance(info, dict):
            self.ws.insert_rows(self.sample_row_idx)
            self.fill_row(self.sample_row_idx, info)
        else:
            raise RuntimeError("info 必须是 dict 或 list of dict")

    def save(self):
        self.wb.save(self.output_path)
        print(f"写入成功，格式已保留，文件已保存到 {self.output_path}")

def main():
    # 示例数据（支持dict或list）
    # 单行模式
    # info = {
    #     "序号": 1,
    #     "证件号": "1234567890",
    #     "姓名": "张三",
    #     "出生年月": "1990-01-01",
    #     "性别": "男",
    #     "民族": "汉",
    #     "考核": {"2023": "合格", "2024": "优秀"}
    # }
    # 多行模式
    info = [
        {
            "序号": 1,
            "证件号": "1234567890",
            "姓名": "张三",
            "出生年月": "1990-01-01",
            "性别": "男",
            "民族": "汉",
            "考核": {"2023": "合格", "2024": "优秀"}
        },
        {
            "序号": 2,
            "证件号": "2234567890",
            "姓名": "李四",
            "出生年月": "1992-03-04",
            "性别": "女",
            "民族": "汉",
            "考核": {"2023": "优秀", "2024": "合格"}
        }
    ]
    template_path = "templates/excel/excel-table-template.xlsx"
    output_path = "output/excel-登记表.xlsx"
    filler = ExcelTemplateFiller(template_path, output_path)
    filler.fill(info)
    filler.save()

if __name__ == "__main__":
    main()